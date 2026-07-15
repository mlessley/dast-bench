from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Protection

from core.models import Confidence, Criterion, CriteriaTaxonomy, ScoreEntry, Vendor, VendorResearchCache, VendorSource
from core.render.stakeholder_workbook import HEADER_ROW, _reviewer_slot_columns, generate_workbook
from core.stakeholder_review import _column_map, merge, populate, snapshot, validate_workbook


def _taxonomy():
    return CriteriaTaxonomy(
        criteria=[
            Criterion(id="c1", category="Coverage", name="Coverage One", description="d", weight=60, rubric="r"),
            Criterion(id="c2", category="DX", name="DX One", description="d", weight=40, rubric="r"),
        ]
    )


def _vendor_with_hands_on_c2():
    vendor = Vendor(id="v1", name="Vendor One", source=VendorSource.DISCOVERED)
    vendor.scores.append(ScoreEntry(criterion_id="c1", score=4.0, evidence="ev1", confidence=Confidence.PAPER))
    vendor.scores.append(ScoreEntry(criterion_id="c2", score=1.0, evidence="paper guess", confidence=Confidence.PAPER))
    vendor.scores.append(ScoreEntry(criterion_id="c2", score=3.5, evidence="hands-on: 7/10 detected", confidence=Confidence.HANDS_ON))
    return vendor


def _generate_with_c2_pending(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy()
    vendor = _vendor_with_hands_on_c2()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={"v1": {"c2"}},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    return out_path


def _row_for(ws, cols, criterion_id):
    crit_col = cols["_criterion_id"]
    for r in range(4, ws.max_row + 1):
        if ws[f"{crit_col}{r}"].value == criterion_id:
            return r
    raise AssertionError(f"row for {criterion_id} not found")


def test_populate_fills_pending_row_and_unlocks_it(tmp_path):
    out_path = _generate_with_c2_pending(tmp_path)
    vendor = _vendor_with_hands_on_c2()
    summary = populate(vendor, out_path)
    assert "populated 1" in summary

    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    crit_id_col = header.index("_criterion_id") + 1
    score_col = header.index("Automated Score") + 1
    evidence_col = header.index("Automated Evidence") + 1
    pending_col = header.index("_pending") + 1
    stakeholder_score_col = header.index("Score") + 1
    stakeholder_dispute_col = header.index("Dispute?") + 1
    stakeholder_rationale_col = header.index("Rationale") + 1
    row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c2")
    assert ws.cell(row=row, column=score_col).value == 3.5
    assert ws.cell(row=row, column=evidence_col).value == "hands-on: 7/10 detected"
    assert ws.cell(row=row, column=pending_col).value == 0
    assert ws.cell(row=row, column=stakeholder_score_col).protection.locked is False
    assert ws.cell(row=row, column=stakeholder_dispute_col).protection.locked is False
    assert ws.cell(row=row, column=stakeholder_rationale_col).protection.locked is False


def test_populate_unlocks_every_reviewer_slot_not_only_the_last(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy()
    vendor = _vendor_with_hands_on_c2()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=3,
        pending_criteria={"v1": {"c2"}},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    populate(vendor, out_path)

    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    crit_id_col = header.index("_criterion_id") + 1
    row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c2")
    for score_col, dispute_col, rationale_col in _reviewer_slot_columns(3):
        assert ws.cell(row=row, column=score_col).protection.locked is False
        assert ws.cell(row=row, column=dispute_col).protection.locked is False
        assert ws.cell(row=row, column=rationale_col).protection.locked is False


def test_populate_leaves_stakeholder_entered_cells_untouched(tmp_path):
    # A stakeholder may pre-fill a Dispute/Rationale note on a PENDING row
    # (e.g. criterion c2) before dast-scan results land. populate() must
    # fill in that same row's Automated Score/Evidence without disturbing
    # the stakeholder-entered cell.
    out_path = _generate_with_c2_pending(tmp_path)
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    crit_id_col = header.index("_criterion_id") + 1
    rationale_col = header.index("Rationale") + 1
    score_col = header.index("Automated Score") + 1
    evidence_col = header.index("Automated Evidence") + 1
    row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c2")
    ws.cell(row=row, column=rationale_col).value = "pre-filled before scan results landed"
    wb = ws.parent
    wb.save(out_path)

    vendor = _vendor_with_hands_on_c2()
    populate(vendor, out_path)

    ws2 = load_workbook(out_path)["v1"]
    assert ws2.cell(row=row, column=rationale_col).value == "pre-filled before scan results landed"
    assert ws2.cell(row=row, column=score_col).value == 3.5
    assert ws2.cell(row=row, column=evidence_col).value == "hands-on: 7/10 detected"


def test_populate_is_a_no_op_when_vendor_has_no_pending_rows(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy()
    vendor = _vendor_with_hands_on_c2()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    summary = populate(vendor, out_path)
    assert "populated 0" in summary


def _generate_two_reviewer_slots(tmp_path, filename="master.xlsx"):
    out_path = tmp_path / filename
    taxonomy = _taxonomy()
    vendor = Vendor(id="v1", name="Vendor One", source=VendorSource.DISCOVERED)
    vendor.scores.append(ScoreEntry(criterion_id="c1", score=4.0, evidence="ev1", confidence=Confidence.PAPER))
    vendor.scores.append(ScoreEntry(criterion_id="c2", score=2.0, evidence="ev2", confidence=Confidence.PAPER))
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=2,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    return out_path


def _slot_letters(slot_number: int) -> tuple[str, str, str]:
    from openpyxl.utils import get_column_letter

    score_col, dispute_col, rationale_col = _reviewer_slot_columns(2)[slot_number - 1]
    return get_column_letter(score_col), get_column_letter(dispute_col), get_column_letter(rationale_col)


def test_merge_fills_blank_master_cells_from_a_valid_returned_copy(tmp_path):
    master_path = _generate_two_reviewer_slots(tmp_path, "master.xlsx")
    copy_path = _generate_two_reviewer_slots(tmp_path, "copy.xlsx")
    score_letter, _, rationale_letter = _slot_letters(1)

    copy_wb = load_workbook(copy_path)
    copy_ws = copy_wb["v1"]
    cols = _column_map(copy_ws)
    row = _row_for(copy_ws, cols, "c1")
    copy_ws[f"{score_letter}{row}"] = 4.5
    copy_ws[f"{rationale_letter}{row}"] = "Confirmed with vendor demo"
    copy_wb.save(copy_path)

    summary = merge(master_path, copy_path)
    assert "merged 1 cell" in summary

    master_ws = load_workbook(master_path)["v1"]
    mcols = _column_map(master_ws)
    mrow = _row_for(master_ws, mcols, "c1")
    assert master_ws[f"{score_letter}{mrow}"].value == 4.5


def test_merge_flags_conflict_without_overwriting(tmp_path):
    master_path = _generate_two_reviewer_slots(tmp_path, "master.xlsx")
    copy_path = _generate_two_reviewer_slots(tmp_path, "copy.xlsx")
    score_letter, _, _ = _slot_letters(1)

    master_wb = load_workbook(master_path)
    master_ws = master_wb["v1"]
    mcols = _column_map(master_ws)
    mrow = _row_for(master_ws, mcols, "c1")
    master_ws[f"{score_letter}{mrow}"] = 3.0
    master_wb.save(master_path)

    copy_wb = load_workbook(copy_path)
    copy_ws = copy_wb["v1"]
    cols = _column_map(copy_ws)
    row = _row_for(copy_ws, cols, "c1")
    copy_ws[f"{score_letter}{row}"] = 4.5
    copy_wb.save(copy_path)

    summary = merge(master_path, copy_path)
    assert "1 conflict" in summary
    assert load_workbook(master_path)["v1"][f"{score_letter}{mrow}"].value == 3.0


def test_merge_addresses_each_slot_independently(tmp_path):
    master_path = _generate_two_reviewer_slots(tmp_path, "master.xlsx")
    copy_path = _generate_two_reviewer_slots(tmp_path, "copy.xlsx")
    slot1_score, _, _ = _slot_letters(1)
    slot2_score, _, _ = _slot_letters(2)

    copy_wb = load_workbook(copy_path)
    copy_ws = copy_wb["v1"]
    cols = _column_map(copy_ws)
    row = _row_for(copy_ws, cols, "c1")
    copy_ws[f"{slot1_score}{row}"] = 4.5
    copy_ws[f"{slot2_score}{row}"] = 3.5
    copy_wb.save(copy_path)

    summary = merge(master_path, copy_path)
    assert "merged 2 cell" in summary

    master_ws = load_workbook(master_path)["v1"]
    mcols = _column_map(master_ws)
    mrow = _row_for(master_ws, mcols, "c1")
    assert master_ws[f"{slot1_score}{mrow}"].value == 4.5
    assert master_ws[f"{slot2_score}{mrow}"].value == 3.5


def test_merge_flags_conflict_on_rationale_alone_when_score_is_blank(tmp_path):
    master_path = _generate_two_reviewer_slots(tmp_path, "master.xlsx")
    copy_path = _generate_two_reviewer_slots(tmp_path, "copy.xlsx")
    _, _, rationale_letter = _slot_letters(1)

    master_wb = load_workbook(master_path)
    master_ws = master_wb["v1"]
    mcols = _column_map(master_ws)
    mrow = _row_for(master_ws, mcols, "c1")
    master_ws[f"{rationale_letter}{mrow}"] = "already had a note"
    master_wb.save(master_path)

    copy_wb = load_workbook(copy_path)
    copy_ws = copy_wb["v1"]
    cols = _column_map(copy_ws)
    row = _row_for(copy_ws, cols, "c1")
    copy_ws[f"{rationale_letter}{row}"] = "a totally different note"
    copy_wb.save(copy_path)

    summary = merge(master_path, copy_path)
    assert "1 conflict" in summary

    master_after = load_workbook(master_path)["v1"]
    assert master_after[f"{rationale_letter}{mrow}"].value == "already had a note"


def test_merge_flags_invalid_score_and_dispute_without_rationale(tmp_path):
    master_path = _generate_two_reviewer_slots(tmp_path, "master.xlsx")
    copy_path = _generate_two_reviewer_slots(tmp_path, "copy.xlsx")
    score_letter, dispute_letter, _ = _slot_letters(1)

    copy_wb = load_workbook(copy_path)
    copy_ws = copy_wb["v1"]
    cols = _column_map(copy_ws)
    row = _row_for(copy_ws, cols, "c1")
    copy_ws[f"{score_letter}{row}"] = 9.0
    row2 = _row_for(copy_ws, cols, "c2")
    copy_ws[f"{dispute_letter}{row2}"] = "Y"
    copy_wb.save(copy_path)

    summary = merge(master_path, copy_path)
    assert "2 invalid" in summary


def test_validate_workbook_flags_dispute_without_rationale_and_invalid_score(tmp_path):
    file_path = _generate_two_reviewer_slots(tmp_path, "review.xlsx")
    slot1_score, slot1_dispute, _ = _slot_letters(1)
    slot2_score, _, _ = _slot_letters(2)

    wb = load_workbook(file_path)
    ws = wb["v1"]
    cols = _column_map(ws)
    row = _row_for(ws, cols, "c1")
    ws[f"{slot1_dispute}{row}"] = "Y"
    row2 = _row_for(ws, cols, "c2")
    ws[f"{slot2_score}{row2}"] = 9.0
    wb.save(file_path)

    issues = validate_workbook(file_path)
    assert len(issues) == 2


def test_validate_workbook_returns_empty_list_for_clean_file(tmp_path):
    file_path = _generate_two_reviewer_slots(tmp_path, "review.xlsx")
    assert validate_workbook(file_path) == []


def test_validate_workbook_flags_pending_row_with_data_entered(tmp_path):
    out_path = _generate_with_c2_pending(tmp_path)
    from openpyxl.utils import get_column_letter

    score_col, _, _ = _reviewer_slot_columns(1)[0]
    score_letter = get_column_letter(score_col)

    wb = load_workbook(out_path)
    ws = wb["v1"]
    cols = _column_map(ws)
    row = _row_for(ws, cols, "c2")
    # Tamper: enter a score on the pending row without touching its lock.
    ws[f"{score_letter}{row}"] = 3.0
    wb.save(out_path)

    issues = validate_workbook(out_path)
    assert any("tampered" in issue for issue in issues)


def test_validate_workbook_flags_pending_row_with_lock_removed(tmp_path):
    out_path = _generate_with_c2_pending(tmp_path)
    from openpyxl.utils import get_column_letter

    score_col, _, _ = _reviewer_slot_columns(1)[0]
    score_letter = get_column_letter(score_col)

    wb = load_workbook(out_path)
    ws = wb["v1"]
    cols = _column_map(ws)
    row = _row_for(ws, cols, "c2")
    # Tamper: remove the lock but leave the cell blank (no data entered).
    ws[f"{score_letter}{row}"].protection = Protection(locked=False)
    wb.save(out_path)

    issues = validate_workbook(out_path)
    assert any("tampered" in issue for issue in issues)


def test_validate_workbook_does_not_flag_untouched_pending_row(tmp_path):
    out_path = _generate_with_c2_pending(tmp_path)
    issues = validate_workbook(out_path)
    assert not any("tampered" in issue for issue in issues)


def test_validate_workbook_flags_pending_row_with_resolved_score_entered(tmp_path):
    out_path = _generate_with_c2_pending(tmp_path)
    wb = load_workbook(out_path)
    ws = wb["v1"]
    cols = _column_map(ws)
    row = _row_for(ws, cols, "c2")
    # Tamper: enter a Resolved Score on the pending row without touching its lock.
    ws[f"{cols['Resolved Score']}{row}"] = 3.0
    wb.save(out_path)

    issues = validate_workbook(out_path)
    assert any("tampered" in issue for issue in issues)


def test_validate_workbook_flags_pending_row_with_resolved_score_lock_removed(tmp_path):
    out_path = _generate_with_c2_pending(tmp_path)
    wb = load_workbook(out_path)
    ws = wb["v1"]
    cols = _column_map(ws)
    row = _row_for(ws, cols, "c2")
    # Tamper: remove the lock on Resolved Score but leave the cell blank.
    ws[f"{cols['Resolved Score']}{row}"].protection = Protection(locked=False)
    wb.save(out_path)

    issues = validate_workbook(out_path)
    assert any("tampered" in issue for issue in issues)


def test_snapshot_copies_file_into_archive_dir(tmp_path):
    file_path = _generate_two_reviewer_slots(tmp_path, "review.xlsx")
    archive_dir = tmp_path / "archive"
    result = snapshot(file_path, "v1", archive_dir, label="baseline")
    assert result.exists()
    assert result.parent == archive_dir
    assert "v1" in result.name
    assert "baseline" in result.name
    assert result.read_bytes() == file_path.read_bytes()


def test_merge_accepts_case_insensitive_dispute_yes_value(tmp_path):
    master_path = _generate_two_reviewer_slots(tmp_path, "master.xlsx")
    copy_path = _generate_two_reviewer_slots(tmp_path, "copy.xlsx")
    _, dispute_letter, rationale_letter = _slot_letters(1)

    copy_wb = load_workbook(copy_path)
    copy_ws = copy_wb["v1"]
    cols = _column_map(copy_ws)
    row = _row_for(copy_ws, cols, "c1")
    copy_ws[f"{dispute_letter}{row}"] = "yes"
    copy_ws[f"{rationale_letter}{row}"] = "Confirmed with vendor demo"
    copy_wb.save(copy_path)

    summary = merge(master_path, copy_path)
    assert "merged 1 cell" in summary

    master_ws = load_workbook(master_path)["v1"]
    mcols = _column_map(master_ws)
    mrow = _row_for(master_ws, mcols, "c1")
    assert master_ws[f"{dispute_letter}{mrow}"].value == "yes"


def test_merge_flags_mixed_case_dispute_without_rationale_as_invalid(tmp_path):
    master_path = _generate_two_reviewer_slots(tmp_path, "master.xlsx")
    copy_path = _generate_two_reviewer_slots(tmp_path, "copy.xlsx")
    _, dispute_letter, _ = _slot_letters(1)

    copy_wb = load_workbook(copy_path)
    copy_ws = copy_wb["v1"]
    cols = _column_map(copy_ws)
    row2 = _row_for(copy_ws, cols, "c2")
    copy_ws[f"{dispute_letter}{row2}"] = "Yes"
    copy_wb.save(copy_path)

    summary = merge(master_path, copy_path)
    assert "1 invalid" in summary


def test_validate_workbook_flags_lowercase_dispute_without_rationale(tmp_path):
    file_path = _generate_two_reviewer_slots(tmp_path, "review.xlsx")
    _, dispute_letter, _ = _slot_letters(1)

    wb = load_workbook(file_path)
    ws = wb["v1"]
    cols = _column_map(ws)
    row = _row_for(ws, cols, "c1")
    ws[f"{dispute_letter}{row}"] = "yes"
    wb.save(file_path)

    issues = validate_workbook(file_path)
    assert len(issues) == 1
    assert "disputed with no rationale" in issues[0]
