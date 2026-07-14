from openpyxl import load_workbook

from core.models import Confidence, Criterion, CriteriaTaxonomy, ScoreEntry, Vendor, VendorSource
from core.render.xlsx import write_xlsx


def _sample_taxonomy() -> CriteriaTaxonomy:
    return CriteriaTaxonomy(
        criteria=[
            Criterion(id="c1", category="Coverage", name="API Coverage", description="d", weight=60, rubric="r"),
            Criterion(id="c2", category="DX", name="Noise Reduction", description="d", weight=40, rubric="r"),
        ]
    )


def _sample_vendor() -> Vendor:
    vendor = Vendor(id="v1", name="Vendor One", source=VendorSource.DISCOVERED)
    vendor.scores.append(ScoreEntry(criterion_id="c1", score=4, evidence="docs", confidence=Confidence.PAPER))
    vendor.scores.append(ScoreEntry(criterion_id="c2", score=2, evidence="trial", confidence=Confidence.HANDS_ON))
    return vendor


def test_write_xlsx_creates_sheet_with_header_and_vendor_column(tmp_path):
    out_path = tmp_path / "comparison-matrix.xlsx"
    write_xlsx(_sample_taxonomy(), [_sample_vendor()], out_path)
    wb = load_workbook(out_path)
    ws = wb["Comparison"]
    header = [cell.value for cell in ws[3]]
    assert header == ["Criterion", "Category", "Weight", "Vendor One"]
    assert ws.cell(row=4, column=4).value == 4
    last_row = [cell.value for cell in ws[ws.max_row]]
    assert last_row[0] == "Weighted Total"
    assert last_row[3] == 4 * 0.6 + 2 * 0.4


def test_write_xlsx_includes_disclaimer(tmp_path):
    out_path = tmp_path / "comparison-matrix.xlsx"
    write_xlsx(_sample_taxonomy(), [_sample_vendor()], out_path)
    wb = load_workbook(out_path)
    ws = wb["Comparison"]
    first_cell = ws.cell(row=1, column=1).value
    assert "Draft/sample output" in first_cell
    assert "not a final vendor recommendation" in first_cell
