from pathlib import Path

from openpyxl import load_workbook

from core.models import (
    Confidence,
    Criterion,
    CriteriaTaxonomy,
    CriterionResearchCache,
    ScoreEntry,
    Vendor,
    VendorResearchCache,
    VendorSource,
)
from openpyxl.utils import get_column_letter

from core.render.stakeholder_workbook import (
    _all_headers,
    _column_index,
    _EXEC_LEGEND_FIRST_ROW,
    _EXEC_LEGEND_HEADER_ROW,
    _EXEC_LEGEND_LINES_TEMPLATE,
    _reviewer_slot_columns,
    _rollup_row_numbers,
    compute_priority_order,
    EXEC_TABLE_FIRST_DATA_ROW,
    EXEC_TABLE_HEADER_ROW,
    generate_workbook,
)


def _taxonomy():
    return CriteriaTaxonomy(
        criteria=[
            Criterion(id="low-weight", category="Cat", name="Low Weight", description="d", weight=5, rubric="r"),
            Criterion(id="high-weight-confident", category="Cat", name="High Confident", description="d", weight=20, rubric="r"),
            Criterion(id="high-weight-shaky", category="Cat", name="High Shaky", description="d", weight=20, rubric="r"),
        ]
    )


def _vendor():
    vendor = Vendor(id="v1", name="V1", source=VendorSource.DISCOVERED)
    vendor.scores.append(ScoreEntry(criterion_id="low-weight", score=5.0, evidence="e", confidence=Confidence.PAPER))
    vendor.scores.append(ScoreEntry(criterion_id="high-weight-confident", score=4.5, evidence="e", confidence=Confidence.PAPER))
    vendor.scores.append(ScoreEntry(criterion_id="high-weight-shaky", score=2.0, evidence="e", confidence=Confidence.PAPER))
    return vendor


def test_priority_order_sorts_by_weight_then_needs_attention():
    taxonomy = _taxonomy()
    vendor = _vendor()
    cache = VendorResearchCache(vendor_id="v1")
    order = compute_priority_order(taxonomy, vendor, cache)
    # both weight-20 criteria outrank the weight-5 one; within the
    # weight-20 band, the low-scoring (<=2.5) one sorts first
    assert order == ["high-weight-shaky", "high-weight-confident", "low-weight"]


def test_priority_order_pulls_up_gap_checked_criteria_even_with_high_score():
    taxonomy = _taxonomy()
    vendor = _vendor()
    vendor.scores[-1] = ScoreEntry(criterion_id="high-weight-shaky", score=4.5, evidence="e", confidence=Confidence.PAPER)
    cache = VendorResearchCache(
        vendor_id="v1",
        criteria={"high-weight-shaky": CriterionResearchCache(reviewed_by_gap_check=True)},
    )
    order = compute_priority_order(taxonomy, vendor, cache)
    assert order == ["high-weight-shaky", "high-weight-confident", "low-weight"]


def _taxonomy_two_criteria():
    return CriteriaTaxonomy(
        criteria=[
            Criterion(id="c1", category="Coverage", name="Coverage One", description="d", weight=60, rubric="r"),
            Criterion(id="c2", category="DX", name="DX One", description="d", weight=40, rubric="r"),
        ]
    )


def _vendor_two_criteria(vendor_id="v1", name="Vendor One"):
    vendor = Vendor(id=vendor_id, name=name, source=VendorSource.DISCOVERED)
    vendor.scores.append(ScoreEntry(criterion_id="c1", score=4.0, evidence="ev1", confidence=Confidence.PAPER))
    vendor.scores.append(ScoreEntry(criterion_id="c2", score=2.0, evidence="ev2", confidence=Confidence.PAPER))
    return vendor


def test_generate_workbook_writes_one_sheet_per_vendor_with_headers(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=2,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    wb = load_workbook(out_path)
    assert wb.sheetnames == ["Executive Summary", "v1"]
    ws = wb["v1"]
    header = [c.value for c in ws[3]]
    assert header[:6] == ["Criterion", "Category", "Weight", "Automated Score", "Automated Evidence", "Automated Confidence"]
    assert header.count("Score") == 2
    assert header.count("Dispute?") == 2
    assert header.count("Rationale") == 2
    assert "Resolved Score" in header
    assert "Resolved By" in header
    assert "Resolved Timestamp" in header
    assert "_criterion_id" in header


def test_generate_workbook_adds_merged_reviewer_slot_group_headers(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=2,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    slot_columns = _reviewer_slot_columns(2)

    slot1_score_col, _, slot1_rationale_col = slot_columns[0]
    slot1_range = f"{get_column_letter(slot1_score_col)}2:{get_column_letter(slot1_rationale_col)}2"
    assert slot1_range in [str(r) for r in ws.merged_cells.ranges]
    slot1_cell = ws.cell(row=2, column=slot1_score_col)
    assert slot1_cell.value == "Reviewer 1 - {name} - {role/title}"
    assert slot1_cell.font.bold is True
    assert slot1_cell.fill.fgColor.rgb == "001F4E78"

    slot2_score_col, _, slot2_rationale_col = slot_columns[1]
    slot2_range = f"{get_column_letter(slot2_score_col)}2:{get_column_letter(slot2_rationale_col)}2"
    assert slot2_range in [str(r) for r in ws.merged_cells.ranges]
    assert ws.cell(row=2, column=slot2_score_col).value == "Reviewer 2 - {name} - {role/title}"


def test_generate_workbook_with_zero_reviewer_slots_has_no_reviewer_columns(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=0,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    assert "Score" not in header
    assert "Dispute?" not in header
    assert "Rationale" not in header
    assert header[:6] == ["Criterion", "Category", "Weight", "Automated Score", "Automated Evidence", "Automated Confidence"]
    assert header[6] == "Resolved Score"


def test_generate_workbook_orders_rows_by_priority_and_fills_automated_data(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    crit_col = header.index("Criterion") + 1
    score_col = header.index("Automated Score") + 1
    crit_id_col = header.index("_criterion_id") + 1
    # c2 (weight 40, score 2.0 <= 2.5) outranks c1 (weight 60, score 4.0)
    # under the priority rule -- wait: weight sorts first, so c1 (60)
    # comes before c2 (40) here since neither ties on weight.
    assert ws.cell(row=4, column=crit_col).value == "Coverage One"
    assert ws.cell(row=4, column=score_col).value == 4.0
    assert ws.cell(row=4, column=crit_id_col).value == "c1"
    assert ws.cell(row=5, column=crit_id_col).value == "c2"


def test_generate_workbook_marks_pending_criteria_with_placeholder_and_no_automated_data(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={"v1": {"c2"}},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    score_col = header.index("Automated Score") + 1
    evidence_col = header.index("Automated Evidence") + 1
    crit_id_col = header.index("_criterion_id") + 1
    pending_col = header.index("_pending") + 1
    row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c2")
    assert ws.cell(row=row, column=pending_col).value == 1
    assert "Pending" in ws.cell(row=row, column=evidence_col).value
    assert ws.cell(row=row, column=score_col).value is None
    non_pending_row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c1")
    assert ws.cell(row=non_pending_row, column=pending_col).value == 0


def test_generate_workbook_adds_score_data_validation_and_locks_pending_rows(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={"v1": {"c2"}},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    assert len(ws.data_validations.dataValidation) >= 1
    score_col = header.index("Score") + 1
    crit_id_col = header.index("_criterion_id") + 1
    pending_row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c2")
    non_pending_row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c1")
    assert ws.cell(row=pending_row, column=score_col).protection.locked is True
    assert ws.cell(row=non_pending_row, column=score_col).protection.locked is False
    assert ws.protection.sheet is True


def test_generate_workbook_writes_provisional_note_above_header(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={"v1": {"c2"}},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    row1 = [c.value for c in ws[1]]
    note = "Provisional — ranking may shift once pending dast-scan results land."
    assert note in row1


def test_generate_workbook_writes_delta_formula_and_partial_completeness_total(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={"v1": {"c2"}},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    delta_col = header.index("Automated vs. Resolved Delta") + 1
    crit_id_col = header.index("_criterion_id") + 1
    row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c1")
    delta_formula = ws.cell(row=row, column=delta_col).value
    assert delta_formula.startswith("=IF(ISBLANK(")

    total_row = ws.cell(row=ws.max_row, column=1).value
    assert total_row == "Weighted Total"
    total_score_cell = ws.cell(row=ws.max_row, column=header.index("Automated Score") + 1).value
    assert total_score_cell.startswith("=")
    assert "available points" in total_score_cell


def test_generate_workbook_applies_column_widths_freeze_panes_and_header_style(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    assert ws.freeze_panes == "G4"
    assert ws.column_dimensions["A"].width == 32  # Criterion

    header = [c.value for c in ws[3]]
    evidence_letter = get_column_letter(header.index("Automated Evidence") + 1)
    assert ws.column_dimensions[evidence_letter].width == 45

    header_cell = ws.cell(row=3, column=1)
    assert header_cell.font.bold is True
    assert header_cell.font.color.rgb == "00FFFFFF"
    assert header_cell.fill.fgColor.rgb == "001F4E78"
    assert header_cell.alignment.wrap_text is True
    assert header_cell.border.top.style == "thin"


def test_generate_workbook_applies_number_formats_banding_border_and_tab_color(tmp_path):
    from core.render.stakeholder_workbook import FIRST_DATA_ROW, _TAB_COLOR_PALETTE

    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
        top_tier_count=0,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    score_col = header.index("Automated Score") + 1
    crit_col = header.index("Criterion") + 1

    assert ws.cell(row=4, column=score_col).number_format == "0.0"
    assert ws.cell(row=4, column=score_col).alignment.horizontal == "right"
    assert ws.cell(row=4, column=crit_col).alignment.horizontal == "left"

    row4_fill = ws.cell(row=4, column=crit_col).fill.fgColor.rgb
    row5_fill = ws.cell(row=5, column=crit_col).fill.fgColor.rgb
    assert row5_fill == "00F2F2F2"
    assert row4_fill != row5_fill

    last_data_row = FIRST_DATA_ROW + len(taxonomy.criteria) - 1
    first_rollup_row = last_data_row + 2
    assert ws.cell(row=first_rollup_row, column=1).border.top.style == "medium"

    assert ws.sheet_properties.tabColor.rgb == "00" + _TAB_COLOR_PALETTE[0]


def test_generate_workbook_adds_dispute_dropdown(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    dispute_col_letter = get_column_letter(header.index("Dispute?") + 1)
    dispute_dvs = [dv for dv in ws.data_validations.dataValidation if dv.formula1 == '"Yes"']
    assert len(dispute_dvs) == 1
    assert f"{dispute_col_letter}4" in str(dispute_dvs[0].sqref)


def test_generate_workbook_adds_executive_summary_sheet_first_with_legend_and_ranked_table(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor_a = Vendor(id="a", name="Vendor A", source=VendorSource.DISCOVERED)
    vendor_a.scores.append(ScoreEntry(criterion_id="c1", score=5.0, evidence="e", confidence=Confidence.PAPER))
    vendor_a.scores.append(ScoreEntry(criterion_id="c2", score=5.0, evidence="e", confidence=Confidence.PAPER))
    vendor_b = Vendor(id="b", name="Vendor B", source=VendorSource.DISCOVERED)
    vendor_b.scores.append(ScoreEntry(criterion_id="c1", score=2.0, evidence="e", confidence=Confidence.PAPER))
    vendor_b.scores.append(ScoreEntry(criterion_id="c2", score=2.0, evidence="e", confidence=Confidence.PAPER))

    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor_b, vendor_a],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={
            "a": VendorResearchCache(vendor_id="a"),
            "b": VendorResearchCache(vendor_id="b"),
        },
        out_path=out_path,
    )

    wb = load_workbook(out_path)
    assert wb.sheetnames[0] == "Executive Summary"
    ws = wb["Executive Summary"]
    assert ws.cell(row=1, column=1).value == "Executive Summary"
    assert ws.cell(row=3, column=1).value == "Legend"
    assert "top 10 priority" in ws.cell(row=5, column=1).value

    header = [c.value for c in ws[EXEC_TABLE_HEADER_ROW]]
    assert header == ["Vendor", "Coverage", "DX", "Weighted Avg Score", "Total Achieved / Available"]

    # Vendor A scored higher on every criterion, so it ranks first
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=1).value == "Vendor A"
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW + 1, column=1).value == "Vendor B"

    category_rows, weighted_total_row = _rollup_row_numbers(taxonomy)
    vendor_headers = _all_headers(1)
    weight_col = get_column_letter(_column_index(vendor_headers, "Weight"))
    evidence_col = get_column_letter(_column_index(vendor_headers, "Automated Evidence"))
    score_col = get_column_letter(_column_index(vendor_headers, "Automated Score"))

    coverage_row = category_rows["Coverage"]
    expected_coverage_formula = (
        f"=IF('a'!{evidence_col}{coverage_row}=0,\"Pending\","
        f"'a'!{weight_col}{coverage_row}/'a'!{evidence_col}{coverage_row}*5)"
    )
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=2).value == expected_coverage_formula

    expected_avg_formula = (
        f"=IF('a'!{evidence_col}{weighted_total_row}=0,\"Pending\","
        f"'a'!{weight_col}{weighted_total_row}/'a'!{evidence_col}{weighted_total_row}*5)"
    )
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=4).value == expected_avg_formula
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=5).value == f"='a'!{score_col}{weighted_total_row}"


def test_generate_workbook_executive_summary_ranks_by_normalized_average_not_raw_achieved_points(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor_x = Vendor(id="x", name="Vendor X", source=VendorSource.DISCOVERED)
    vendor_x.scores.append(ScoreEntry(criterion_id="c1", score=5.0, evidence="e", confidence=Confidence.PAPER))
    vendor_x.scores.append(ScoreEntry(criterion_id="c2", score=1.0, evidence="e", confidence=Confidence.PAPER))
    vendor_y = Vendor(id="y", name="Vendor Y", source=VendorSource.DISCOVERED)
    vendor_y.scores.append(ScoreEntry(criterion_id="c1", score=4.0, evidence="e", confidence=Confidence.PAPER))
    vendor_y.scores.append(ScoreEntry(criterion_id="c2", score=4.0, evidence="e", confidence=Confidence.PAPER))

    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor_y, vendor_x],
        reviewer_slots=1,
        pending_criteria={"x": {"c2"}},
        research_caches={
            "x": VendorResearchCache(vendor_id="x"),
            "y": VendorResearchCache(vendor_id="y"),
        },
        out_path=out_path,
    )

    ws = load_workbook(out_path)["Executive Summary"]
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=1).value == "Vendor X"
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW + 1, column=1).value == "Vendor Y"


def test_generate_workbook_executive_summary_sorts_all_pending_vendor_last(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor_a = Vendor(id="a", name="Vendor A", source=VendorSource.DISCOVERED)
    vendor_a.scores.append(ScoreEntry(criterion_id="c1", score=1.0, evidence="e", confidence=Confidence.PAPER))
    vendor_a.scores.append(ScoreEntry(criterion_id="c2", score=1.0, evidence="e", confidence=Confidence.PAPER))
    vendor_b = Vendor(id="b", name="Vendor B", source=VendorSource.DISCOVERED)
    vendor_b.scores.append(ScoreEntry(criterion_id="c1", score=5.0, evidence="e", confidence=Confidence.PAPER))
    vendor_b.scores.append(ScoreEntry(criterion_id="c2", score=5.0, evidence="e", confidence=Confidence.PAPER))

    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor_a, vendor_b],
        reviewer_slots=1,
        pending_criteria={"b": {"c1", "c2"}},
        research_caches={
            "a": VendorResearchCache(vendor_id="a"),
            "b": VendorResearchCache(vendor_id="b"),
        },
        out_path=out_path,
    )

    ws = load_workbook(out_path)["Executive Summary"]
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=1).value == "Vendor A"
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW + 1, column=1).value == "Vendor B"


def test_generate_workbook_executive_summary_includes_bar_chart(tmp_path):
    from openpyxl.chart import BarChart

    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["Executive Summary"]
    assert len(ws._charts) == 1
    assert isinstance(ws._charts[0], BarChart)


def test_generate_workbook_freeze_panes_include_weight_column(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    assert ws.freeze_panes == "G4"


def test_generate_workbook_wraps_evidence_and_rationale_text(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    evidence_col = header.index("Automated Evidence") + 1
    rationale_col = header.index("Rationale") + 1
    score_col = header.index("Automated Score") + 1

    assert ws.cell(row=4, column=evidence_col).alignment.wrap_text is True
    assert ws.cell(row=4, column=rationale_col).alignment.wrap_text is True
    assert ws.cell(row=4, column=score_col).alignment.wrap_text is not True
    assert ws.row_dimensions[4].height is None


def test_generate_workbook_bands_continuously_under_tier_highlight(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = CriteriaTaxonomy(
        criteria=[
            Criterion(id="c1", category="Cat", name="C1", description="d", weight=100, rubric="r"),
            Criterion(id="c2", category="Cat", name="C2", description="d", weight=90, rubric="r"),
            Criterion(id="c3", category="Cat", name="C3", description="d", weight=80, rubric="r"),
            Criterion(id="c4", category="Cat", name="C4", description="d", weight=70, rubric="r"),
        ]
    )
    vendor = Vendor(id="v1", name="V1", source=VendorSource.DISCOVERED)
    for criterion_id in ("c1", "c2", "c3", "c4"):
        vendor.scores.append(ScoreEntry(criterion_id=criterion_id, score=4.0, evidence="e", confidence=Confidence.PAPER))

    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
        top_tier_count=2,
    )
    ws = load_workbook(out_path)["v1"]
    assert ws.cell(row=4, column=1).fill.fgColor.rgb == "00FFF2CC"  # tier, even (i=0)
    assert ws.cell(row=5, column=1).fill.fgColor.rgb == "00F9E79F"  # tier, odd (i=1)
    assert ws.cell(row=6, column=1).fill.fgColor.rgb == "00000000"  # non-tier, even (i=2) -- no fill
    assert ws.cell(row=7, column=1).fill.fgColor.rgb == "00F2F2F2"  # non-tier, odd (i=3) -- banded


def test_generate_workbook_executive_summary_legend_has_border_and_fill(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["Executive Summary"]
    legend_last_row = _EXEC_LEGEND_FIRST_ROW + len(_EXEC_LEGEND_LINES_TEMPLATE) - 1

    top_left = ws.cell(row=_EXEC_LEGEND_HEADER_ROW, column=1)
    assert top_left.fill.fgColor.rgb == "00F2F2F2"
    assert top_left.border.top.style == "thin"

    bottom_right = ws.cell(row=legend_last_row, column=5)
    assert bottom_right.fill.fgColor.rgb == "00F2F2F2"
    assert bottom_right.border.bottom.style == "thin"


def test_generate_workbook_executive_summary_highlights_top_vendor_row(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor_a = Vendor(id="a", name="Vendor A", source=VendorSource.DISCOVERED)
    vendor_a.scores.append(ScoreEntry(criterion_id="c1", score=5.0, evidence="e", confidence=Confidence.PAPER))
    vendor_a.scores.append(ScoreEntry(criterion_id="c2", score=5.0, evidence="e", confidence=Confidence.PAPER))
    vendor_b = Vendor(id="b", name="Vendor B", source=VendorSource.DISCOVERED)
    vendor_b.scores.append(ScoreEntry(criterion_id="c1", score=2.0, evidence="e", confidence=Confidence.PAPER))
    vendor_b.scores.append(ScoreEntry(criterion_id="c2", score=2.0, evidence="e", confidence=Confidence.PAPER))

    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor_b, vendor_a],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={
            "a": VendorResearchCache(vendor_id="a"),
            "b": VendorResearchCache(vendor_id="b"),
        },
        out_path=out_path,
    )

    ws = load_workbook(out_path)["Executive Summary"]
    assert ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=1).value == "Vendor A"
    top_row_cell = ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=1)
    assert top_row_cell.font.bold is True
    assert top_row_cell.fill.fgColor.rgb == "00FFF2CC"

    second_row_cell = ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW + 1, column=1)
    assert second_row_cell.value == "Vendor B"
    assert second_row_cell.font.bold is not True
    assert second_row_cell.fill.fgColor.rgb == "00000000"


def test_generate_workbook_writes_updated_pending_row_scope_text(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={"v1": {"c2"}},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    header = [c.value for c in ws[3]]
    evidence_col = header.index("Automated Evidence") + 1
    crit_id_col = header.index("_criterion_id") + 1
    row = next(r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=crit_id_col).value == "c2")
    assert ws.cell(row=row, column=evidence_col).value == (
        "Pending — dast-scan results not yet available. "
        "Do not score or edit this row; it will be populated in Round 2."
    )


def test_generate_workbook_rollup_block_has_summary_header(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["v1"]
    category_rows, _ = _rollup_row_numbers(taxonomy)
    summary_header_row = min(category_rows.values()) - 1

    cell = ws.cell(row=summary_header_row, column=1)
    assert cell.value == "Summary"
    assert cell.font.bold is True
    assert cell.fill.fgColor.rgb == "001F4E78"


def test_generate_workbook_executive_summary_legend_explains_weight(tmp_path):
    out_path = tmp_path / "review.xlsx"
    taxonomy = _taxonomy_two_criteria()
    vendor = _vendor_two_criteria()
    generate_workbook(
        taxonomy=taxonomy,
        vendors=[vendor],
        reviewer_slots=1,
        pending_criteria={},
        research_caches={"v1": VendorResearchCache(vendor_id="v1")},
        out_path=out_path,
    )
    ws = load_workbook(out_path)["Executive Summary"]
    weight_legend_row = _EXEC_LEGEND_FIRST_ROW + 5
    legend_text = ws.cell(row=weight_legend_row, column=1).value
    assert "Weight" in legend_text
    assert "set by the evaluator" in legend_text
