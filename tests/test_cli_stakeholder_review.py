import yaml
from openpyxl import load_workbook
from typer.testing import CliRunner

from core.cli import app

runner = CliRunner()


def _setup_repo(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    (tmp_path / "data" / "candidates").mkdir(parents=True)
    (tmp_path / "data" / "research-cache").mkdir(parents=True)
    criteria = {
        "version": 1,
        "criteria": [
            {"id": "c1", "category": "Coverage", "name": "Coverage One", "description": "d", "weight": 100.0, "rubric": "r"},
        ],
    }
    (tmp_path / "data" / "criteria.yaml").write_text(yaml.safe_dump(criteria))
    vendor = {
        "id": "v1",
        "name": "Vendor One",
        "source": "discovered",
        "status": "finalist",
        "scores": [{"criterion_id": "c1", "score": 4.0, "evidence": "ev1", "confidence": "paper", "timestamp": "2026-01-01T00:00:00"}],
        "hands_on_results": [],
        "observations": [],
    }
    (tmp_path / "data" / "candidates" / "v1.yaml").write_text(yaml.safe_dump(vendor))


def test_cli_stakeholder_review_generate_creates_workbook(tmp_path, monkeypatch):
    _setup_repo(tmp_path, monkeypatch)
    out_path = tmp_path / "review.xlsx"
    result = runner.invoke(
        app,
        [
            "stakeholder-review", "generate",
            "--vendor-id", "v1",
            "--stakeholder", ":DAST SME",
            "--out", str(out_path),
        ],
    )
    assert result.exit_code == 0, result.output
    wb = load_workbook(out_path)
    assert "v1" in wb.sheetnames


def test_cli_stakeholder_review_generate_fails_on_unscored_criterion(tmp_path, monkeypatch):
    _setup_repo(tmp_path, monkeypatch)
    extra_criteria = yaml.safe_load((tmp_path / "data" / "criteria.yaml").read_text())
    extra_criteria["criteria"].append(
        {"id": "c2", "category": "DX", "name": "DX One", "description": "d", "weight": 0.0, "rubric": "r"}
    )
    (tmp_path / "data" / "criteria.yaml").write_text(yaml.safe_dump(extra_criteria))
    result = runner.invoke(
        app,
        [
            "stakeholder-review", "generate",
            "--vendor-id", "v1",
            "--stakeholder", ":DAST SME",
            "--out", str(tmp_path / "review.xlsx"),
        ],
    )
    assert result.exit_code == 1
    assert "error:" in result.output
