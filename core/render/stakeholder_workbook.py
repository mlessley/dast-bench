from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..models import CriteriaTaxonomy, Vendor, VendorResearchCache
from .markdown import _ordered_categories

SCORE_VALUES = [1.0, 1.5, 2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0]
_PENDING_TEXT = (
    "Pending — dast-scan results not yet available. "
    "Do not score or edit this row; it will be populated in Round 2."
)
_BASE_HEADERS = [
    "Criterion",
    "Category",
    "Weight",
    "Automated Score",
    "Automated Evidence",
    "Automated Confidence",
]
_RESOLUTION_HEADERS = ["Resolved Score", "Resolved By", "Resolved Timestamp", "Automated vs. Resolved Delta"]
_HIDDEN_HEADERS = ["_criterion_id", "_pending", "_effective_score"]

_TIER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_UNFILLED_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

_HEADER_FILL_COLOR = "1F4E78"
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color=_HEADER_FILL_COLOR, end_color=_HEADER_FILL_COLOR, fill_type="solid")
_HEADER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_FIXED_COLUMN_WIDTHS = {
    "Criterion": 32,
    "Category": 16,
    "Weight": 10,
    "Automated Score": 14,
    "Automated Evidence": 45,
    "Automated Confidence": 16,
    "Resolved Score": 14,
    "Resolved By": 16,
    "Resolved Timestamp": 18,
    "Automated vs. Resolved Delta": 14,
}
_COLUMN_WIDTHS_BY_SUFFIX = [
    ("Rationale", 40),
    ("Dispute?", 12),
    ("Score", 12),
]


def _column_width_for(header: str) -> float | None:
    if header in _FIXED_COLUMN_WIDTHS:
        return _FIXED_COLUMN_WIDTHS[header]
    for suffix, width in _COLUMN_WIDTHS_BY_SUFFIX:
        if header.endswith(suffix):
            return width
    return None


_BAND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_ROLLUP_BORDER = Border(top=Side(style="medium"))

_RIGHT_ALIGN_NUMERIC_HEADERS_EXACT = {"Weight", "Automated vs. Resolved Delta"}
_RIGHT_ALIGN_NUMERIC_HEADERS_SUFFIX = "Score"

_TAB_COLOR_PALETTE = ["2E86AB", "A23B72", "F18F01", "C73E1D", "3B1F2B", "6A994E"]


def _tab_color_for(index: int) -> str:
    return _TAB_COLOR_PALETTE[index % len(_TAB_COLOR_PALETTE)]


_EXEC_SHEET_NAME = "Overview"
_EXEC_TAB_COLOR = "1F4E78"
_EXEC_TITLE_ROW = 1
_EXEC_ABOUT_HEADER_ROW = 3
_EXEC_ABOUT_FIRST_ROW = 4
_EXEC_ABOUT_LINES = [
    "This workbook ranks DAST (dynamic application security testing) vendors against a weighted set "
    "of evaluation criteria, blending automated desk research with hands-on benchmark testing where "
    "available.",
    "Reviewers can weigh in on any criterion on that vendor's tab and resolve disagreements with the "
    "automated score; resolved scores take precedence in the totals below. Claim a reviewer slot once "
    "on the \"Reviewers\" tab and it applies automatically to every vendor tab.",
    "See the \"How This Works\" tab (last tab) for the full methodology, scoring formulas, and "
    "definitions.",
]

EXEC_TABLE_HEADER_ROW = _EXEC_ABOUT_FIRST_ROW + len(_EXEC_ABOUT_LINES) + 1
EXEC_TABLE_FIRST_DATA_ROW = EXEC_TABLE_HEADER_ROW + 1

_HOW_IT_WORKS_SHEET_NAME = "How This Works"
_HOW_IT_WORKS_TAB_COLOR = "808080"
_METHODOLOGY_HEADER_ROW = 3
_METHODOLOGY_FIRST_ROW = 4
_METHODOLOGY_LINES = [
    "The evaluation criteria, categories, and weights were defined up front, before any vendor was "
    "researched, based on standard DAST-buying considerations — coverage, detection quality, "
    "production safety, developer experience, reporting, and deployment/data governance. Fixing the "
    "rubric first keeps the comparison consistent instead of retrofitting criteria to favor a "
    "particular tool.",
    "Each criterion has a written rubric describing what a 1 vs. a 3 vs. a 5 looks like. Automated "
    "scores come from structured desk research against that rubric — vendor docs, pricing pages, "
    "release notes, and public third-party reviews — with the evidence and reasoning captured "
    "alongside every score so it's traceable, not a black box.",
    "Two criteria, Detection Accuracy and False Positive Rate, also get a hands-on pass: the tool is "
    "run against known-vulnerable benchmark targets (OWASP Juice Shop, VAmPI) and the desk-research "
    "score is replaced with one based on real scan output. Confidence is upgraded from 'paper' to "
    "'hands-on' when that happens.",
    "Reviewer input on each vendor's tab exists to catch what desk research misses or gets wrong — "
    "resolved scores from reviewers take precedence over the automated ones in the final totals.",
]

_LEGEND_HEADER_ROW = _METHODOLOGY_FIRST_ROW + len(_METHODOLOGY_LINES) + 1
_LEGEND_FIRST_ROW = _LEGEND_HEADER_ROW + 1
_LEGEND_LINES = [
    "Pending rows: dast-scan results not yet available; locked until Round 2 populate fills them in. "
    "Until then, treat the Overview ranking as provisional.",
    "Tier highlight (pink): top {top_tier_count} priority Score cells are tinted while still empty, "
    "as a reminder they still need a value.",
    "Dispute = Yes requires a non-blank Rationale; discuss unresolved disputes before finalizing a score.",
    "Automated Confidence: 'paper' = desk research only; 'hands-on' = verified via dast-scan.",
    "Weighted Avg Score is normalized to a 0-5 scale and is comparable across vendors even when the "
    "number of pending criteria differs. Row order below is fixed when this workbook is generated and "
    "does not auto-resort if scores change later.",
    "Weight (on each vendor sheet) is each criterion's relative importance in the overall weighted "
    "score — the same value used in the actual score calculation, set by the evaluator when the "
    "criteria taxonomy was defined, not something a reviewer sets.",
]


def _rollup_row_numbers(taxonomy: CriteriaTaxonomy) -> tuple[dict[str, int], int]:
    last_data_row = FIRST_DATA_ROW + len(taxonomy.criteria) - 1
    categories = _ordered_categories(taxonomy)
    first_category_row = last_data_row + 2
    category_rows = {category: first_category_row + i for i, category in enumerate(categories)}
    weighted_total_row = first_category_row + len(categories)
    return category_rows, weighted_total_row


def _weighted_avg_score(taxonomy: CriteriaTaxonomy, vendor: Vendor, pending_for_vendor: set[str]) -> float | None:
    achieved = 0.0
    available = 0.0
    for criterion in taxonomy.criteria:
        if criterion.id in pending_for_vendor:
            continue
        entry = vendor.score_for(criterion.id)
        score = entry.score if entry else 0.0
        achieved += criterion.weight * score
        available += criterion.weight
    if available == 0:
        return None
    return achieved / available


def _add_executive_summary_sheet(
    wb: Workbook,
    taxonomy: CriteriaTaxonomy,
    vendors: list[Vendor],
    pending_criteria: dict[str, set[str]],
    headers: list[str],
) -> None:
    ws = wb.create_sheet(title=_EXEC_SHEET_NAME)
    ws.sheet_properties.tabColor = _EXEC_TAB_COLOR

    title_cell = ws.cell(row=_EXEC_TITLE_ROW, column=1, value=_EXEC_SHEET_NAME)
    title_cell.font = Font(bold=True, size=16)
    ws.row_dimensions[_EXEC_TITLE_ROW].height = 28

    about_header_cell = ws.cell(row=_EXEC_ABOUT_HEADER_ROW, column=1, value="About This Report")
    about_header_cell.font = Font(bold=True)
    ws.row_dimensions[_EXEC_ABOUT_HEADER_ROW].height = 20
    for i, line in enumerate(_EXEC_ABOUT_LINES):
        ws.cell(row=_EXEC_ABOUT_FIRST_ROW + i, column=1, value=line)

    categories = _ordered_categories(taxonomy)
    category_rows, weighted_total_row = _rollup_row_numbers(taxonomy)
    table_headers = ["Vendor"] + categories + ["Weighted Avg Score", "Total Achieved / Available"]
    avg_col = 2 + len(categories)

    for col_idx, header_name in enumerate(table_headers, start=1):
        cell = ws.cell(row=EXEC_TABLE_HEADER_ROW, column=col_idx, value=header_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[EXEC_TABLE_HEADER_ROW].height = 34

    weight_col_letter = get_column_letter(_column_index(headers, "Weight"))
    evidence_col_letter = get_column_letter(_column_index(headers, "Automated Evidence"))
    score_col_letter = get_column_letter(_column_index(headers, "Automated Score"))

    def _sort_key(vendor: Vendor) -> tuple[bool, float]:
        avg = _weighted_avg_score(taxonomy, vendor, pending_criteria.get(vendor.id, set()))
        return (avg is None, -(avg or 0.0))

    ranked_vendors = sorted(vendors, key=_sort_key)

    for i, vendor in enumerate(ranked_vendors):
        row_num = EXEC_TABLE_FIRST_DATA_ROW + i
        sheet_name = vendor.id[:31]
        ws.cell(row=row_num, column=1, value=vendor.name)

        for col_offset, category in enumerate(categories):
            r = category_rows[category]
            formula = (
                f"=IF('{sheet_name}'!{evidence_col_letter}{r}=0,\"Pending\","
                f"'{sheet_name}'!{weight_col_letter}{r}/'{sheet_name}'!{evidence_col_letter}{r}*5)"
            )
            cell = ws.cell(row=row_num, column=2 + col_offset, value=formula)
            cell.number_format = "0.0"
            cell.alignment = Alignment(horizontal="right")

        avg_formula = (
            f"=IF('{sheet_name}'!{evidence_col_letter}{weighted_total_row}=0,\"Pending\","
            f"'{sheet_name}'!{weight_col_letter}{weighted_total_row}/"
            f"'{sheet_name}'!{evidence_col_letter}{weighted_total_row}*5)"
        )
        avg_cell = ws.cell(row=row_num, column=avg_col, value=avg_formula)
        avg_cell.number_format = "0.0"
        avg_cell.alignment = Alignment(horizontal="right")

        ws.cell(
            row=row_num,
            column=avg_col + 1,
            value=f"='{sheet_name}'!{score_col_letter}{weighted_total_row}",
        )

    for col_idx, header_name in enumerate(table_headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 40 if header_name == "Vendor" else 20

    if ranked_vendors:
        for col in range(1, len(table_headers) + 1):
            top_cell = ws.cell(row=EXEC_TABLE_FIRST_DATA_ROW, column=col)
            top_cell.font = Font(bold=True)
            top_cell.fill = _TIER_FILL

    ws.protection.sheet = True

    if ranked_vendors:
        last_row = EXEC_TABLE_FIRST_DATA_ROW + len(ranked_vendors) - 1
        chart = BarChart()
        chart.type = "col"
        chart.title = "Weighted Avg Score by Vendor"
        chart.x_axis.title = "Vendor"
        chart.y_axis.title = "Weighted Avg Score (0-5)"
        data = Reference(ws, min_col=avg_col, min_row=EXEC_TABLE_HEADER_ROW, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=EXEC_TABLE_FIRST_DATA_ROW, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{last_row + 3}")


def _is_right_aligned_numeric(header_name: str) -> bool:
    return (
        header_name in _RIGHT_ALIGN_NUMERIC_HEADERS_EXACT
        or header_name.endswith(_RIGHT_ALIGN_NUMERIC_HEADERS_SUFFIX)
    )


def _is_wrapped_text(header_name: str) -> bool:
    return header_name == "Automated Evidence" or header_name.endswith("Rationale")


HEADER_ROW = 3
FIRST_DATA_ROW = 4


def compute_priority_order(
    taxonomy: CriteriaTaxonomy, vendor: Vendor, research_cache: VendorResearchCache
) -> list[str]:
    def sort_key(criterion):
        entry = vendor.score_for(criterion.id)
        score = entry.score if entry else 0.0
        cache_entry = research_cache.criteria.get(criterion.id)
        gap_checked = cache_entry.reviewed_by_gap_check if cache_entry else False
        needs_attention = score <= 2.5 or gap_checked
        return (-criterion.weight, 0 if needs_attention else 1, criterion.id)

    return [c.id for c in sorted(taxonomy.criteria, key=sort_key)]


_REVIEWER_SLOT_SUB_HEADERS = ["Score", "Dispute?", "Rationale"]


def _reviewer_slot_headers(reviewer_slots: int) -> list[str]:
    return _REVIEWER_SLOT_SUB_HEADERS * reviewer_slots


def _reviewer_slot_columns(reviewer_slots: int) -> list[tuple[int, int, int]]:
    base = len(_BASE_HEADERS)
    return [
        (base + i * 3 + 1, base + i * 3 + 2, base + i * 3 + 3)
        for i in range(reviewer_slots)
    ]


def _reviewer_slot_count_from_headers(headers: list[str]) -> int:
    count = 0
    idx = len(_BASE_HEADERS)
    while idx + 2 < len(headers) and headers[idx:idx + 3] == _REVIEWER_SLOT_SUB_HEADERS:
        count += 1
        idx += 3
    return count


_REVIEWERS_SHEET_NAME = "Reviewers"
_REVIEWERS_TAB_COLOR = "8E44AD"
_REVIEWERS_TITLE_ROW = 1
_REVIEWERS_INSTRUCTIONS_ROW = 3
_REVIEWERS_HEADER_ROW = 5
_REVIEWERS_FIRST_DATA_ROW = 6
_REVIEWERS_SLOT_COL = 1
_REVIEWERS_NAME_COL = 2
_REVIEWERS_ROLE_COL = 3


def _unclaimed_reviewer_label(slot_number: int) -> str:
    return f"Reviewer {slot_number} - {{name}} - {{role/title}}"


def _reviewer_slot_group_header_formula(slot_index: int, slot_number: int) -> str:
    """A vendor sheet's reviewer-slot header pulls live from the Reviewers sheet, so
    claiming a slot there propagates to every vendor tab instead of being re-typed per tab."""
    row = _REVIEWERS_FIRST_DATA_ROW + slot_index
    name_ref = f"Reviewers!{get_column_letter(_REVIEWERS_NAME_COL)}{row}"
    role_ref = f"Reviewers!{get_column_letter(_REVIEWERS_ROLE_COL)}{row}"
    placeholder = _unclaimed_reviewer_label(slot_number)
    return f'=IF({name_ref}="","{placeholder}",{name_ref}&" - "&{role_ref})'


def _write_reviewer_slot_group_headers(ws, reviewer_slots: int) -> None:
    for slot_index, (score_col, _dispute_col, rationale_col) in enumerate(_reviewer_slot_columns(reviewer_slots)):
        ws.merge_cells(start_row=2, start_column=score_col, end_row=2, end_column=rationale_col)
        anchor = ws.cell(
            row=2, column=score_col,
            value=_reviewer_slot_group_header_formula(slot_index, slot_index + 1),
        )
        anchor.font = _HEADER_FONT
        anchor.fill = _HEADER_FILL
        anchor.border = _HEADER_BORDER
        anchor.alignment = _HEADER_ALIGNMENT
        anchor.protection = Protection(locked=True)


def _add_reviewers_sheet(wb: Workbook, reviewer_slots: int) -> None:
    ws = wb.create_sheet(title=_REVIEWERS_SHEET_NAME)
    ws.sheet_properties.tabColor = _REVIEWERS_TAB_COLOR

    title_cell = ws.cell(row=_REVIEWERS_TITLE_ROW, column=1, value=_REVIEWERS_SHEET_NAME)
    title_cell.font = Font(bold=True, size=14)
    ws.row_dimensions[_REVIEWERS_TITLE_ROW].height = 26

    ws.cell(
        row=_REVIEWERS_INSTRUCTIONS_ROW,
        column=1,
        value="Claim a slot by filling in your name and role below — it appears automatically on every vendor tab.",
    )

    header_labels = ["Slot", "Name", "Role / Title"]
    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=_REVIEWERS_HEADER_ROW, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _HEADER_BORDER
        cell.alignment = _HEADER_ALIGNMENT
    ws.row_dimensions[_REVIEWERS_HEADER_ROW].height = 20

    for i in range(reviewer_slots):
        row = _REVIEWERS_FIRST_DATA_ROW + i
        if i % 2 == 1:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = _BAND_FILL

        slot_cell = ws.cell(row=row, column=_REVIEWERS_SLOT_COL, value=i + 1)
        slot_cell.alignment = Alignment(horizontal="center")

        for col in (_REVIEWERS_NAME_COL, _REVIEWERS_ROLE_COL):
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.column_dimensions[get_column_letter(_REVIEWERS_SLOT_COL)].width = 8
    ws.column_dimensions[get_column_letter(_REVIEWERS_NAME_COL)].width = 28
    ws.column_dimensions[get_column_letter(_REVIEWERS_ROLE_COL)].width = 28
    ws.protection.sheet = True


def _add_how_it_works_sheet(wb: Workbook, top_tier_count: int) -> None:
    ws = wb.create_sheet(title=_HOW_IT_WORKS_SHEET_NAME)
    ws.sheet_properties.tabColor = _HOW_IT_WORKS_TAB_COLOR
    ws.column_dimensions["A"].width = 100

    title_cell = ws.cell(row=1, column=1, value=_HOW_IT_WORKS_SHEET_NAME)
    title_cell.font = Font(bold=True, size=16)
    ws.row_dimensions[1].height = 28

    methodology_header_cell = ws.cell(row=_METHODOLOGY_HEADER_ROW, column=1, value="Methodology")
    methodology_header_cell.font = Font(bold=True)
    ws.row_dimensions[_METHODOLOGY_HEADER_ROW].height = 20
    for i, line in enumerate(_METHODOLOGY_LINES):
        cell = ws.cell(row=_METHODOLOGY_FIRST_ROW + i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    legend_header_cell = ws.cell(row=_LEGEND_HEADER_ROW, column=1, value="Legend")
    legend_header_cell.font = Font(bold=True)
    ws.row_dimensions[_LEGEND_HEADER_ROW].height = 20
    for i, line in enumerate(_LEGEND_LINES):
        cell = ws.cell(row=_LEGEND_FIRST_ROW + i, column=1, value=line.format(top_tier_count=top_tier_count))
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    legend_last_row = _LEGEND_FIRST_ROW + len(_LEGEND_LINES) - 1
    for row in range(_LEGEND_FIRST_ROW, legend_last_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = _BAND_FILL
            cell.border = _HEADER_BORDER
            if col == 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _all_headers(reviewer_slots: int) -> list[str]:
    return _BASE_HEADERS + _reviewer_slot_headers(reviewer_slots) + _RESOLUTION_HEADERS + _HIDDEN_HEADERS


def _column_index(headers: list[str], name: str) -> int:
    return headers.index(name) + 1


def generate_workbook(
    taxonomy: CriteriaTaxonomy,
    vendors: list[Vendor],
    reviewer_slots: int,
    pending_criteria: dict[str, set[str]],
    research_caches: dict[str, VendorResearchCache],
    out_path: Path,
    top_tier_count: int = 10,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    headers = _all_headers(reviewer_slots)
    _add_executive_summary_sheet(wb, taxonomy, vendors, pending_criteria, headers)
    _add_reviewers_sheet(wb, reviewer_slots)
    for vendor_index, vendor in enumerate(vendors):
        ws = wb.create_sheet(title=vendor.id[:31])
        ws.sheet_properties.tabColor = _tab_color_for(vendor_index)
        ws.append(["Provisional — ranking may shift once pending dast-scan results land."])
        ws.append([])
        ws.append(headers)
        ws.freeze_panes = "G4"
        for col_idx, header_name in enumerate(headers, start=1):
            width = _column_width_for(header_name)
            if width is not None:
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.border = _HEADER_BORDER
            cell.alignment = _HEADER_ALIGNMENT
        ws.row_dimensions[HEADER_ROW].height = 36
        ws.row_dimensions[2].height = 24
        _write_reviewer_slot_group_headers(ws, reviewer_slots)
        pending_for_vendor = pending_criteria.get(vendor.id, set())
        cache = research_caches.get(vendor.id) or VendorResearchCache(vendor_id=vendor.id)
        order = compute_priority_order(taxonomy, vendor, cache)

        slot_columns = _reviewer_slot_columns(reviewer_slots)
        score_cols = [score_col for score_col, _, _ in slot_columns] + [_column_index(headers, "Resolved Score")]

        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(str(v) for v in SCORE_VALUES) + '"',
            allow_blank=True,
        )
        ws.add_data_validation(dv)

        dispute_cols = [dispute_col for _, dispute_col, _ in slot_columns]
        dispute_dv = DataValidation(type="list", formula1='"Yes"', allow_blank=True)
        ws.add_data_validation(dispute_dv)

        for i, criterion_id in enumerate(order):
            row_num = FIRST_DATA_ROW + i
            criterion = taxonomy.get(criterion_id)
            entry = vendor.score_for(criterion_id)
            is_pending = criterion_id in pending_for_vendor
            row = [criterion.name, criterion.category, criterion.weight]
            if is_pending:
                row += [None, _PENDING_TEXT, None]
            else:
                row += [entry.score if entry else None, entry.evidence if entry else None, entry.confidence.value if entry else None]
            row += [None] * len(_reviewer_slot_headers(reviewer_slots))
            row += [None, None, None, None]
            row += [criterion_id, 1 if is_pending else 0, None]
            ws.append(row)

            if i % 2 == 1:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = _BAND_FILL

            for col_idx, header_name in enumerate(headers, start=1):
                if header_name in _HIDDEN_HEADERS:
                    continue
                cell = ws.cell(row=row_num, column=col_idx)
                if _is_right_aligned_numeric(header_name):
                    cell.number_format = "0.0"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=_is_wrapped_text(header_name))

            for col in score_cols:
                cell = ws.cell(row=row_num, column=col)
                cell.protection = Protection(locked=is_pending)
                dv.add(cell)
                if i < top_tier_count:
                    ws.conditional_formatting.add(
                        cell.coordinate,
                        CellIsRule(operator="equal", formula=['""'], fill=_UNFILLED_FILL),
                    )

            for col in dispute_cols:
                dispute_dv.add(ws.cell(row=row_num, column=col))

            editable_non_score_cols = [
                col for _, dispute_col, rationale_col in slot_columns for col in (dispute_col, rationale_col)
            ] + [_column_index(headers, h) for h in ("Resolved By", "Resolved Timestamp")]
            for col in editable_non_score_cols:
                ws.cell(row=row_num, column=col).protection = Protection(locked=is_pending)

        resolved_col_letter = get_column_letter(_column_index(headers, "Resolved Score"))
        automated_col_letter = get_column_letter(_column_index(headers, "Automated Score"))
        effective_col_letter = get_column_letter(_column_index(headers, "_effective_score"))
        pending_col_letter = get_column_letter(_column_index(headers, "_pending"))
        weight_col_letter = get_column_letter(_column_index(headers, "Weight"))
        category_col_letter = get_column_letter(_column_index(headers, "Category"))
        delta_col = _column_index(headers, "Automated vs. Resolved Delta")
        effective_col = _column_index(headers, "_effective_score")
        last_data_row = FIRST_DATA_ROW + len(order) - 1

        for row_num in range(FIRST_DATA_ROW, last_data_row + 1):
            resolved_ref = f"{resolved_col_letter}{row_num}"
            automated_ref = f"{automated_col_letter}{row_num}"
            ws.cell(row=row_num, column=effective_col).value = f"=IF(ISBLANK({resolved_ref}),{automated_ref},{resolved_ref})"
            ws.cell(row=row_num, column=delta_col).value = f'=IF(ISBLANK({resolved_ref}),"",{resolved_ref}-{automated_ref})'

        def _points_formulas(category_filter: str | None) -> tuple[str, str]:
            weight_range = f"{weight_col_letter}{FIRST_DATA_ROW}:{weight_col_letter}{last_data_row}"
            effective_range = f"{effective_col_letter}{FIRST_DATA_ROW}:{effective_col_letter}{last_data_row}"
            pending_range = f"{pending_col_letter}{FIRST_DATA_ROW}:{pending_col_letter}{last_data_row}"
            if category_filter is None:
                achieved = f"=SUMPRODUCT({weight_range},{effective_range},(1-{pending_range}))/5"
                available = f"=SUMPRODUCT({weight_range},(1-{pending_range}))"
            else:
                category_range = f"{category_col_letter}{FIRST_DATA_ROW}:{category_col_letter}{last_data_row}"
                achieved = f'=SUMPRODUCT(({category_range}=\"{category_filter}\")*{weight_range}*{effective_range}*(1-{pending_range}))/5'
                available = f'=SUMPRODUCT(({category_range}=\"{category_filter}\")*{weight_range}*(1-{pending_range}))'
            return achieved, available

        weight_header_col = _column_index(headers, "Weight")
        evidence_header_col = _column_index(headers, "Automated Evidence")
        score_header_col = _column_index(headers, "Automated Score")

        def _write_rollup_row(label: str, category_filter: str | None) -> None:
            ws.append([label] + [None] * (len(headers) - 1))
            r = ws.max_row
            achieved_formula, available_formula = _points_formulas(category_filter)
            ws.cell(row=r, column=weight_header_col).value = achieved_formula
            ws.cell(row=r, column=evidence_header_col).value = available_formula
            weight_ref = f"{weight_col_letter}{r}"
            evidence_ref = f"{get_column_letter(evidence_header_col)}{r}"
            ws.cell(row=r, column=score_header_col).value = (
                f'=TEXT({weight_ref},"0.0")&"/"&TEXT({evidence_ref},"0")&" available points"'
            )

        ws.append(["Summary"] + [None] * (len(headers) - 1))
        summary_header_row = ws.max_row
        summary_cell = ws.cell(row=summary_header_row, column=1)
        summary_cell.font = _HEADER_FONT
        ws.row_dimensions[summary_header_row].height = 20
        summary_cell.fill = _HEADER_FILL
        category_rows_for_border, _ = _rollup_row_numbers(taxonomy)
        first_rollup_row = category_rows_for_border[_ordered_categories(taxonomy)[0]]
        for category in _ordered_categories(taxonomy):
            _write_rollup_row(category, category)
        _write_rollup_row("Weighted Total", None)
        for col in range(1, len(headers) + 1):
            ws.cell(row=first_rollup_row, column=col).border = _ROLLUP_BORDER

        for hidden_name in _HIDDEN_HEADERS:
            ws.column_dimensions[get_column_letter(_column_index(headers, hidden_name))].hidden = True
        ws.protection.sheet = True
    _add_how_it_works_sheet(wb, top_tier_count)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
