from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter

from .models import Vendor
from .render.stakeholder_workbook import (
    FIRST_DATA_ROW,
    HEADER_ROW,
    SCORE_VALUES,
    _REVIEWERS_FIRST_DATA_ROW,
    _REVIEWERS_NAME_COL,
    _REVIEWERS_SHEET_NAME,
    _reviewer_slot_columns,
    _reviewer_slot_count_from_headers,
)

_DISPUTE_YES_VALUES = {"y", "yes"}


def _is_dispute_yes(value) -> bool:
    return isinstance(value, str) and value.strip().lower() in _DISPUTE_YES_VALUES


def _column_map(ws) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for cell in ws[HEADER_ROW]:
        if cell.value:
            mapping[cell.value] = cell.column_letter
    return mapping


def _slot_letters(ws) -> list[tuple[str, str, str]]:
    headers = [c.value for c in ws[HEADER_ROW]]
    reviewer_slots = _reviewer_slot_count_from_headers(headers)
    return [
        (get_column_letter(score_col), get_column_letter(dispute_col), get_column_letter(rationale_col))
        for score_col, dispute_col, rationale_col in _reviewer_slot_columns(reviewer_slots)
    ]


def populate(vendor: Vendor, file_path: Path) -> str:
    wb = load_workbook(file_path)
    sheet_name = vendor.id[:31]
    if sheet_name not in wb.sheetnames:
        return f"error: vendor '{vendor.id}' has no sheet in {file_path}"
    ws = wb[sheet_name]
    cols = _column_map(ws)
    crit_id_col = cols["_criterion_id"]
    pending_col = cols["_pending"]
    score_col = cols["Automated Score"]
    evidence_col = cols["Automated Evidence"]
    confidence_col = cols["Automated Confidence"]
    slot_letters = _slot_letters(ws)

    filled = 0
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        criterion_id = ws[f"{crit_id_col}{row}"].value
        if not criterion_id:
            continue
        if ws[f"{pending_col}{row}"].value != 1:
            continue
        entry = vendor.score_for(criterion_id)
        if entry is None:
            continue
        ws[f"{score_col}{row}"] = entry.score
        ws[f"{evidence_col}{row}"] = entry.evidence
        ws[f"{confidence_col}{row}"] = entry.confidence.value
        ws[f"{pending_col}{row}"] = 0
        for score_letter, dispute_letter, rationale_letter in slot_letters:
            for letter in (score_letter, dispute_letter, rationale_letter):
                ws[f"{letter}{row}"].protection = Protection(locked=False)
        filled += 1

    wb.save(file_path)
    return f"populated {filled} pending row(s) for '{vendor.id}'"


def _conflicts(existing, incoming) -> bool:
    return existing is not None and incoming is not None and existing != incoming


def merge(master_path: Path, from_path: Path) -> str:
    master_wb = load_workbook(master_path)
    from_wb = load_workbook(from_path)

    merged = 0
    invalid = 0
    conflicts = 0

    for sheet_name in master_wb.sheetnames:
        if sheet_name not in from_wb.sheetnames:
            continue
        m_ws = master_wb[sheet_name]
        f_ws = from_wb[sheet_name]
        m_cols = _column_map(m_ws)
        f_cols = _column_map(f_ws)

        if "_criterion_id" not in m_cols or "_criterion_id" not in f_cols:
            # Not a per-vendor rollup sheet (e.g. the Executive Summary tab).
            continue

        m_crit_col = m_cols["_criterion_id"]
        f_crit_col = f_cols["_criterion_id"]
        m_pending_col = m_cols["_pending"]

        m_row_by_crit = {
            m_ws[f"{m_crit_col}{r}"].value: r
            for r in range(FIRST_DATA_ROW, m_ws.max_row + 1)
            if m_ws[f"{m_crit_col}{r}"].value
        }
        f_row_by_crit = {
            f_ws[f"{f_crit_col}{r}"].value: r
            for r in range(FIRST_DATA_ROW, f_ws.max_row + 1)
            if f_ws[f"{f_crit_col}{r}"].value
        }

        m_slot_letters = _slot_letters(m_ws)
        f_slot_letters = _slot_letters(f_ws)
        shared_slot_count = min(len(m_slot_letters), len(f_slot_letters))

        for slot_index in range(shared_slot_count):
            score_letter, dispute_letter, rationale_letter = m_slot_letters[slot_index]
            f_score_letter, f_dispute_letter, f_rationale_letter = f_slot_letters[slot_index]
            for criterion_id, f_row in f_row_by_crit.items():
                f_score = f_ws[f"{f_score_letter}{f_row}"].value
                f_dispute = f_ws[f"{f_dispute_letter}{f_row}"].value
                f_rationale = f_ws[f"{f_rationale_letter}{f_row}"].value
                if f_score is None and f_dispute is None and f_rationale is None:
                    continue
                m_row = m_row_by_crit.get(criterion_id)
                if m_row is None or m_ws[f"{m_pending_col}{m_row}"].value == 1:
                    continue
                valid = (f_score is None or f_score in SCORE_VALUES) and (
                    not _is_dispute_yes(f_dispute) or bool(f_rationale)
                )
                if not valid:
                    invalid += 1
                    continue

                existing_score = m_ws[f"{score_letter}{m_row}"].value
                existing_dispute = m_ws[f"{dispute_letter}{m_row}"].value
                existing_rationale = m_ws[f"{rationale_letter}{m_row}"].value
                if (
                    _conflicts(existing_score, f_score)
                    or _conflicts(existing_dispute, f_dispute)
                    or _conflicts(existing_rationale, f_rationale)
                ):
                    conflicts += 1
                    continue
                m_ws[f"{score_letter}{m_row}"] = f_score
                m_ws[f"{dispute_letter}{m_row}"] = f_dispute
                m_ws[f"{rationale_letter}{m_row}"] = f_rationale
                merged += 1

    master_wb.save(master_path)
    return f"merged {merged} cell(s), {invalid} invalid, {conflicts} conflict(s)"


def validate_workbook(file_path: Path) -> list[str]:
    wb = load_workbook(file_path)
    reviewers_ws = wb[_REVIEWERS_SHEET_NAME] if _REVIEWERS_SHEET_NAME in wb.sheetnames else None
    issues: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cols = _column_map(ws)
        if "_criterion_id" not in cols:
            # Not a per-vendor rollup sheet (e.g. the Executive Summary or Reviewers tab).
            continue
        crit_col = cols["_criterion_id"]
        slot_letters = _slot_letters(ws)

        for slot_number, (score_letter, dispute_letter, rationale_letter) in enumerate(slot_letters, start=1):
            # A slot's identity is claimed on the Reviewers sheet (Name column), not on
            # the vendor tab -- the vendor-tab header is a formula that mirrors it.
            slot_claimed = bool(
                reviewers_ws
                and reviewers_ws.cell(row=_REVIEWERS_FIRST_DATA_ROW + slot_number - 1, column=_REVIEWERS_NAME_COL).value
            )
            for row in range(FIRST_DATA_ROW, ws.max_row + 1):
                criterion_id = ws[f"{crit_col}{row}"].value
                if not criterion_id:
                    continue
                score = ws[f"{score_letter}{row}"].value
                dispute = ws[f"{dispute_letter}{row}"].value
                rationale = ws[f"{rationale_letter}{row}"].value
                if score is not None and score not in SCORE_VALUES:
                    issues.append(f"{sheet_name}/{criterion_id}: Reviewer {slot_number} score {score!r} is not a valid value")
                if _is_dispute_yes(dispute) and not rationale:
                    issues.append(f"{sheet_name}/{criterion_id}: Reviewer {slot_number} disputed with no rationale")
                if not slot_claimed and any(v is not None for v in (score, dispute, rationale)):
                    issues.append(
                        f"{sheet_name}/{criterion_id}: Reviewer {slot_number} has responses but "
                        "the slot is unclaimed (header still shows placeholder text)"
                    )
        resolved_h = "Resolved Score"
        by_h = "Resolved By"
        ts_h = "Resolved Timestamp"
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            criterion_id = ws[f"{crit_col}{row}"].value
            if not criterion_id:
                continue
            resolved = ws[f"{cols[resolved_h]}{row}"].value
            resolved_by = ws[f"{cols[by_h]}{row}"].value
            resolved_ts = ws[f"{cols[ts_h]}{row}"].value
            if resolved is not None and not (resolved_by and resolved_ts):
                issues.append(f"{sheet_name}/{criterion_id}: resolved score present without Resolved By/Timestamp")

        pending_col = cols["_pending"]
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            criterion_id = ws[f"{crit_col}{row}"].value
            if not criterion_id:
                continue
            if ws[f"{pending_col}{row}"].value != 1:
                continue
            tampered = False
            for score_letter, dispute_letter, rationale_letter in slot_letters:
                for letter in (score_letter, dispute_letter, rationale_letter):
                    cell = ws[f"{letter}{row}"]
                    if cell.value is not None or cell.protection.locked is False:
                        tampered = True
                        break
                if tampered:
                    break
            if not tampered:
                for header in (resolved_h, by_h, ts_h):
                    cell = ws[f"{cols[header]}{row}"]
                    if cell.value is not None or cell.protection.locked is False:
                        tampered = True
                        break
            if tampered:
                issues.append(
                    f"{sheet_name}/{criterion_id}: pending row has been tampered with "
                    "(lock removed or data entered before Round 2)"
                )
    return issues


def snapshot(file_path: Path, vendor_id: str, archive_dir: Path, label: str | None = None) -> Path:
    archive_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"-{label}" if label else ""
    dest = archive_dir / f"{date.today().isoformat()}-{vendor_id}{suffix}.xlsx"
    shutil.copy2(file_path, dest)
    return dest
